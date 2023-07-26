import time
import pandas as pd
import requests
from lxml import etree
import urllib.parse
from openpyxl.styles import PatternFill
from openpyxl import Workbook

item_list = []
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36'}


def main_page():
    # Make a single request to fetch all the pages
    url = "https://www.juzikong.com/authors/87c50259-8e62-4868-ba5f-d20f4d67da67?page="
    base_url = "https://www.juzikong.com"

    for i in range(1, 3):
        page_url = url + str(i)
        result = requests.get(page_url, headers=headers).content.decode()
        html = etree.HTML(result)
        div_list = html.xpath(".//div[contains(@class,'list')]/section")

        for section in div_list:
            relative_url = section.xpath(".//div[contains(@class,'content')]/a/@href")[0]
            nickname = section.xpath(".//a[contains(@class,'nickname')]/text()")[0].strip()
            content = section.xpath(".//div[contains(@class,'content')]//span/text()")[0]
            full_url = urllib.parse.urljoin(base_url, relative_url)
            comment_count = section.xpath(".//a[contains(@class,'comment')]/span/text()")[0]
            like_count = section.xpath(".//span[contains(@class,'like')]/span/text()")[0]
            item = {'昵称': nickname, 'content': content, 'source': full_url, 'comment': comment_count,
                    'like': like_count}
            item_list.append(item)
        print("正在爬取第{}页".format(i))
        time.sleep(0.1)

    df = pd.DataFrame(item_list)

    # Set style for even and odd rows
    even_fill = PatternFill(fill_type='solid', fgColor="FFFFCC")
    odd_fill = PatternFill(fill_type='solid', fgColor="CCFFCC")

    # Create Excel workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Sheet1'
    worksheet.append(list(df.columns))  # 添加列名到工作表
    # Write data to worksheet
    for row_data in df.values:
        worksheet.append(list(row_data))

    # Apply fill color to even and odd rows
    for row in worksheet.iter_rows(min_row=1):
        row_num = row[0].row
        if row_num % 2 == 0:
            for cell in row:
                cell.fill = even_fill
        else:
            for cell in row:
                cell.fill = odd_fill

    # Save workbook
    workbook.save('D:/Document/Python/{}经典语录.xlsx'.format('模仿鲁迅'))
    workbook.close()


def main():
    main_page()


if __name__ == '__main__':
    main()
