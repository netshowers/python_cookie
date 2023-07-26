import time
import pandas as pd
import requests
from lxml import etree
import urllib.parse
from openpyxl.styles import PatternFill


item_list = []
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36'}


def main_page():
    # Make a single request to fetch all the pages
    url = "https://www.juzikong.com/authors/87c50259-8e62-4868-ba5f-d20f4d67da67?page="
    base_url = "https://www.juzikong.com"

    for i in range(1, 11):
        page_url = url + str(i)
        result = requests.get(page_url, headers=headers).content.decode()
        html = etree.HTML(result)
        div_list = html.xpath(".//div[contains(@class,'list')]/section")

        for section in div_list:
            relative_url = section.xpath(".//div[contains(@class,'content')]/a/@href")[0]
            nickname = section.xpath(".//a[contains(@class,'nickname')]/text()")[0].strip()
            content = section.xpath(".//div[contains(@class,'content')]//span/text()")[0]
            full_url = urllib.parse.urljoin(base_url, relative_url)
            comment_count = section.xpath(".//a[contains(@class,'comment')]/span/text()")[0]
            like_count = section.xpath(".//span[contains(@class,'like')]/span/text()")[0]
            item = {'name': nickname, 'content': content, 'source': full_url, 'comment': comment_count,
                    'like': like_count}
            item_list.append(item)
        print("正在爬取第{}页".format(i))
        time.sleep(0.1)

    df = pd.DataFrame(item_list)
    # Create an Excel writer using openpyxl
    writer = pd.ExcelWriter('D:/Document/Python/{}经典语录_0.xlsx'.format('模仿鲁迅'), engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Get the workbook and the sheet
    workbook = writer.book
    worksheet = workbook['Sheet1']

    # Define fill colors for even and odd rows
    even_fill = PatternFill(fill_type='solid', fgColor="FFFF99")
    odd_fill = PatternFill(fill_type='solid', fgColor="99FF99")

    # Apply fill color to rows
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=5, values_only=True):
        if row[0] is not None and isinstance(row[0], str):
            row_num = worksheet.cell(row=worksheet.max_row, column=1).row
            if row_num % 2 == 0:
                for cell in worksheet[row_num]:
                    cell.fill = even_fill
            else:
                for cell in worksheet[row_num]:
                    cell.fill = odd_fill

    workbook.save('D:/Document/Python/{}经典语录_1.xlsx'.format('模仿鲁迅'))
    writer.close()


def main():
    main_page()


if __name__ == '__main__':
    main()
